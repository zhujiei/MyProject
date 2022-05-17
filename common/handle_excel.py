import openpyxl
from common.handle_path import TESTCASSE_DIR


class HandleExcel:

    def __init__(self, filename, sheetname):
        '''

        :param filename: excel文件名字
        :param sheetname: 表单名
        '''
        self.filename = filename
        self.sheetname = sheetname

    def read_data(self):
        # 加载excel工作薄
        workbook = openpyxl.load_workbook(self.filename)
        # 选中表单
        sh = workbook[self.sheetname]
        #获取数据
        cel = list(sh.rows)
        # 变量第一行数据
        tital = [i.value for i in cel[0]]
        cases = []
        # 遍历除第一行以外的行
        for item in cel[1:]:
            # 获取该行的数据
            data = [i.value for i in item]
            # 把第一行数据和当前行数据打包成字典
            dic = dict(zip(tital, data))
            # 把字典添加到列表中
            cases.append(dic)
        return cases

    def write_data(self, row, column, value):
        '''
        :param row: 写入行
        :param column: 写入列
        :param value: 写入值
        :return:
        '''
        workbook = openpyxl.load_workbook(self.filename)
        sh = workbook[self.sheetname]
        sh.cell(row=row, column=column, value=value)
        workbook.save(self.filename)

    def close_file(self):
        self.close_file(self.filename)

if __name__ == '__main__':
    t = HandleExcel(TESTCASSE_DIR+"\人群管理用例.xlsx", "人群")
    t1 = t.write_data(row=2, column=12, value='nihao')
    print(t1)
    print(type(t1))
